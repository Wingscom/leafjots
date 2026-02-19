"""ExcelWriter — builds bangketoan.xlsx with openpyxl."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from cryptotax.report.data_collector import ReportData

# Sheet definitions: (sheet_name, headers, data_attr, number_formats)
# number_formats: dict of column_index (0-based) → openpyxl number format
SHEET_DEFS: list[tuple[str, list[str], str, dict[int, str]]] = [
    (
        "summary",
        ["Metric", "Value"],
        "summary",
        {},
    ),
    (
        "balance_sheet_by_qty",
        ["Account Type", "Subtype", "Symbol", "Label", "Quantity"],
        "balance_sheet_qty",
        {4: "#,##0.000000"},
    ),
    (
        "balance_sheet_by_value_USD",
        ["Account Type", "Subtype", "Symbol", "Label", "Value (USD)"],
        "balance_sheet_usd",
        {4: "$#,##0.00"},
    ),
    (
        "balance_sheet_by_value_VND",
        ["Account Type", "Subtype", "Symbol", "Label", "Value (VND)"],
        "balance_sheet_vnd",
        {4: "#,##0"},
    ),
    (
        "income_statement",
        ["Account Type", "Symbol", "Label", "Value (USD)", "Value (VND)"],
        "income_statement",
        {3: "$#,##0.00", 4: "#,##0"},
    ),
    (
        "flows_by_qty",
        ["Timestamp", "Type", "Description", "Symbol", "Quantity"],
        "flows_qty",
        {4: "#,##0.000000"},
    ),
    (
        "flows_by_value_USD",
        ["Timestamp", "Type", "Description", "Symbol", "Value (USD)"],
        "flows_usd",
        {4: "$#,##0.00"},
    ),
    (
        "realized_gains",
        ["Symbol", "Quantity", "Cost Basis (USD)", "Proceeds (USD)", "Gain/Loss (USD)", "Holding Days", "Buy Date", "Sell Date"],
        "realized_gains",
        {1: "#,##0.000000", 2: "$#,##0.00", 3: "$#,##0.00", 4: "$#,##0.00"},
    ),
    (
        "open_lots",
        ["Symbol", "Remaining Qty", "Cost Basis/Unit (USD)", "Total Cost (USD)", "Buy Date"],
        "open_lots",
        {1: "#,##0.000000", 2: "$#,##0.00", 3: "$#,##0.00"},
    ),
    (
        "journal",
        ["Timestamp", "Type", "Description", "Account Type", "Symbol", "Label", "Quantity", "Value (USD)", "Value (VND)"],
        "journal",
        {6: "#,##0.000000", 7: "$#,##0.00", 8: "#,##0"},
    ),
    (
        "tax_summary",
        ["Timestamp", "Symbol", "Quantity", "Value (VND)", "Tax (VND)", "Status"],
        "tax_summary",
        {2: "#,##0.000000", 3: "#,##0", 4: "#,##0"},
    ),
    (
        "warnings",
        ["Warning"],
        "warnings",
        {},
    ),
    (
        "wallets",
        ["Chain", "Address", "Label", "Sync Status"],
        "wallets",
        {},
    ),
    (
        "settings",
        ["Setting", "Value"],
        "settings_data",
        {},
    ),
]

HEADER_FONT = Font(bold=True)


class ExcelWriter:
    """Writes ReportData to an in-memory Excel buffer."""

    def write_to_buffer(self, data: ReportData) -> BytesIO:
        """Create a complete bangketoan.xlsx workbook and return as BytesIO."""
        wb = Workbook()

        for idx, (sheet_name, headers, data_attr, num_fmts) in enumerate(SHEET_DEFS):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Header row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = HEADER_FONT

            # Data rows
            sheet_data = getattr(data, data_attr, [])

            if data_attr == "warnings":
                # Warnings is a list of strings, not tuples
                for row_idx, warning in enumerate(sheet_data, start=2):
                    ws.cell(row=row_idx, column=1, value=warning)
            else:
                for row_idx, row in enumerate(sheet_data, start=2):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # Apply number format if defined
                        fmt = num_fmts.get(col_idx - 1)  # col_idx is 1-based, num_fmts keys are 0-based
                        if fmt:
                            cell.number_format = fmt

            # Auto-fit column widths (approximate)
            _auto_fit_columns(ws)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


def _auto_fit_columns(ws) -> None:
    """Set column widths based on content (approximate)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        # Add padding, cap at 50
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
