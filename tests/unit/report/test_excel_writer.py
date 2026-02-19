"""Tests for ExcelWriter — openpyxl workbook generation."""

from io import BytesIO

from openpyxl import load_workbook

from cryptotax.report.data_collector import ReportData
from cryptotax.report.excel_writer import ExcelWriter, SHEET_DEFS


class TestExcelWriterEmpty:
    def test_produces_valid_xlsx(self):
        data = ReportData()
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        assert isinstance(buf, BytesIO)
        assert buf.tell() == 0  # Rewound to start
        assert len(buf.getvalue()) > 0

    def test_has_14_sheets(self):
        data = ReportData()
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        assert len(wb.sheetnames) == 14

    def test_sheet_names_match_spec(self):
        data = ReportData()
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        expected = [sd[0] for sd in SHEET_DEFS]
        assert wb.sheetnames == expected

    def test_each_sheet_has_headers(self):
        data = ReportData()
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        for sheet_name, headers, _, _ in SHEET_DEFS:
            ws = wb[sheet_name]
            row1 = [cell.value for cell in ws[1]]
            assert row1[:len(headers)] == headers, f"Sheet '{sheet_name}' headers mismatch"


class TestExcelWriterWithData:
    def test_summary_sheet_rows(self):
        data = ReportData(
            summary=[
                ("Entity", "test"),
                ("Period Start", "2025-01-01"),
                ("Total Realized Gain (USD)", 1000.0),
            ]
        )
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        ws = wb["summary"]
        assert ws.cell(row=2, column=1).value == "Entity"
        assert ws.cell(row=2, column=2).value == "test"
        assert ws.cell(row=4, column=2).value == 1000.0

    def test_realized_gains_data(self):
        data = ReportData(
            realized_gains=[
                ("ETH", 1.0, 2000.0, 3000.0, 1000.0, 151, "2025-01-01", "2025-06-01"),
            ]
        )
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        ws = wb["realized_gains"]
        assert ws.cell(row=2, column=1).value == "ETH"
        assert ws.cell(row=2, column=2).value == 1.0
        assert ws.cell(row=2, column=5).value == 1000.0
        assert ws.cell(row=2, column=6).value == 151

    def test_warnings_as_strings(self):
        data = ReportData(
            warnings=["Missing price: Buy ETH at 2025-01-01", "Unbalanced qty: gas (delta=0.01)"]
        )
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        ws = wb["warnings"]
        assert ws.cell(row=2, column=1).value == "Missing price: Buy ETH at 2025-01-01"
        assert ws.cell(row=3, column=1).value == "Unbalanced qty: gas (delta=0.01)"

    def test_multiple_sheets_with_data(self):
        data = ReportData(
            balance_sheet_qty=[("ASSET", "native_asset", "ETH", "ETH", 1.5)],
            open_lots=[("ETH", 0.5, 2000.0, 1000.0, "2025-01-01")],
            wallets=[("ethereum", "0xabc", "Main Wallet", "IDLE")],
            settings_data=[("Entity", "test"), ("FIFO Method", "GLOBAL_FIFO")],
        )
        writer = ExcelWriter()
        buf = writer.write_to_buffer(data)

        wb = load_workbook(buf)
        # Balance sheet
        ws = wb["balance_sheet_by_qty"]
        assert ws.cell(row=2, column=3).value == "ETH"
        assert ws.cell(row=2, column=5).value == 1.5

        # Wallets
        ws = wb["wallets"]
        assert ws.cell(row=2, column=1).value == "ethereum"
        assert ws.cell(row=2, column=2).value == "0xabc"
